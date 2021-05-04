import pandas as pd
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


header = pd.read_csv(sys.argv[1], sep='\t',index_col=False, names=['x'],header=None).iloc[:,0]
names=list(header)
names.insert(0, '')

df = pd.read_table(sys.argv[1], names=names, skiprows=0, index_col=0)
df = pd.DataFrame(df)
df.fillna(0,inplace=True)

file=sys.argv[1]

a=file.split('.')

out=a[0]+"_matrix.csv"
outx=a[0]+"_matrix.xls"
outxx=a[0]+"_matrix.xlsx"

df.to_csv(out,sep='\t')
#this below write column up to 256 columns
#df.to_excel(outx)

wb = Workbook()

ws = wb.active

for r in dataframe_to_rows(df, index=True, header=True):
    ws.append(r)

for cell in ws['A'] + ws[1]:
    cell.style = 'Pandas'

wb.save(outxx)


