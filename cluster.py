from scipy.cluster.hierarchy import ward, fcluster
from scipy.cluster.hierarchy import dendrogram, linkage
from scipy.spatial.distance import squareform
import pandas as pd
import sys
import numpy as np
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows


def cluster_df(df, method='single', threshold=5):
    '''
    Accepts a square distance matrix as an indexed DataFrame and returns a dict of index keyed flat clusters 
    Performs single linkage clustering by default, see scipy.cluster.hierarchy.linkage docs for others
    '''

    dm_cnd = squareform(df.values)
    clusters1 = fcluster(linkage(dm_cnd,
                                method=method,
                                metric='precomputed'),
                        criterion='distance',
                        t=threshold)
    names_clusters = {s:c for s, c in zip(df.columns, clusters1)}
    return names_clusters

#sys.argv[1] is the .matrix file from MTBseq in Groups
file=sys.argv[1]

header = pd.read_csv(file, sep='\t',index_col=False, names=['x'],header=None).iloc[:,0]
names=list(header)
names.insert(0, '')



df = pd.read_table(file, delim_whitespace=True,names=names, skiprows=0, index_col=0)
df = pd.DataFrame(df)
df.fillna(0,inplace=True)



D=np.triu(df.T,1) + df

threshold=sys.argv[2]

a=cluster_df(D,method='single',threshold=threshold)

a_df=pd.DataFrame.from_dict(a,orient='index',columns=['Clu'])

#a_df.to_csv('clusters5.csv')

d = defaultdict(list)
for x in a:
    d[a[x]].append(x)

for j in d:
    if (len(d[j]) >= 1):
        #d[j].append(len(d[j]))
        d[j].insert(0,len(d[j]))
        print(j,d[j])
#        print j,len(d[j])
d_df=pd.DataFrame.from_dict(d,orient='index')

#d_df.to_csv('cluster_449_linkage_cutoff-5.csv')


a=file.split('.')

out0=a[0]+"_cluster-"+threshold+".csv"
out1=a[0]+"_cluster_element-"+threshold+".csv"

a_df.to_csv(out0,sep='\t')

d_df.to_csv(out1,sep='\t')



#outxx=a[0]+"_matrix.xlsx"
#wb = Workbook()

#ws = wb.active

#for r in dataframe_to_rows(df, index=True, header=True):
#    ws.append(r)

#for cell in ws['A'] + ws[1]:
#    cell.style = 'Pandas'

#wb.save(outxx)








